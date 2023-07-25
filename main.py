# IMPORT
from jira import JIRA
from jira.resources import User
import openpyxl
import configparser
import os
import time

# TODO add a column for backlogs on in sprint
# FUNCTIONS
def is_excel_recent():
    _input = "n"
    if os.path.getmtime(EXCEL_LOCATION) < time.time() - 600:
        print("___________________________")
        print("THE EXCEL IS OLD, YOU SURELY")
        print("FORGOT TO COPY IT FROM CLOUD")
        print("PRESS Y TO CONTINUE")
        while _input != "y":
            _input = input()
        print("___________________________")


def get_issue_hyperlink(issue_id):
    return f"https://sgbcn.atlassian.net/browse/{str(issue_id)}"


def get_user_id(email):
    # http://pythonjira.com/how-to-set-an-assignee-on-jira-with-python/
    params = {
        "query": email,
        "includeActive": True,
        "includeInactive": False,
    }
    list_search = jira._fetch_pages(User, None, "user/search", params=params)
    return list_search[0].accountId


def create_jira(project, issuetype, summary, assignee=None, description=None, parent=None, customfield=None, priority=None, estimate=None, reporter=None, components=None):
    fields = {"project": {"key": project}, "issuetype": {"name": issuetype}, "summary": summary}
    if assignee is not None:
        jira_user_id = get_user_id(assignee)
        fields["assignee"] = {"accountId": jira_user_id}
    if description is not None:
        fields["description"] = description
    if parent is not None:
        fields["parent"] = {"key": parent}
    if components is not None:
        fields["components"] = [{"add": comp} for comp in components]
    if customfield is not None:
        fields[f"customfield_{customfield['number']}"] = customfield['content']
    if priority is not None:
        fields["priority"] = {"name": priority.capitalize()}
    if reporter is not None:
        reporter_jira_user_id = get_user_id(reporter)
        fields["reporter"] = {"accountId": reporter_jira_user_id}

    print(fields)
    _issue = jira.create_issue(fields=fields)
    if estimate is not None:
        _issue.update(fields={"timetracking": {"originalEstimate": estimate}})
    return _issue


def first_time_config():
    config = configparser.ConfigParser()
    config['EXCEL'] = {"location": r"D:\Jira.xlsx", "spritesheet name": "new", "summary column": "F", "issuetype column": "A", "key column": "B", "assignee column": "G", "priority column": "E", "description column": "I", "estimate column": "J", "validation color": ["9", "FFA9D08E"], "start row": 5}
    config['JIRA'] = {"email": "user.name@domain.tld", "token": "get yours at https://id.atlassian.com/manage-profile/security/api-tokens", "jira server": "https://atlassian.com", "project": "project name"}

    with open(_CONFIG_FILE_LOCATION, 'w') as configfile:
        config.write(configfile)
    _input = "n"
    print("___________________________")
    print("CONFIGURE YOUR SETTINGS.INI")
    print("___________________________")
    print("DONT FORGET TO SAVE THE INI")
    print("___________________________")
    print("PRESS Y TO CONTINUE")
    while _input != "y":
        _input = input()
    print("___________________________")


# CONSTANT
_CONFIG_FILE_LOCATION = os.path.join(os.path.dirname(__file__), "settings.ini")

# First time config
if not os.path.isfile(_CONFIG_FILE_LOCATION):
    first_time_config()

_config = configparser.ConfigParser()
_config.read(_CONFIG_FILE_LOCATION)

EMAIL = _config['JIRA']["email"]
TOKEN = _config['JIRA']["token"]
JIRA_SERVER = _config['JIRA']["jira server"]
PROJECT = _config['JIRA']["project"]
PREFIX = _config['JIRA']["prefix"]
EXCEL_LOCATION = _config['EXCEL']["location"]
SPRITESHEET_NAME = _config['EXCEL']["spritesheet name"]
SUMMARY_COLUMN = _config['EXCEL']["summary column"]
DESCRIPTION_COLUMN = _config['EXCEL']["description column"]
KEY_COLUMN = _config['EXCEL']["key column"]
ISSUETYPE_COLUMN = _config['EXCEL']["issuetype column"]
ASSIGNEE_COLUMN = _config['EXCEL']["assignee column"]
PRIORITY_COLUMN = _config['EXCEL']["priority column"]
ESTIMATE_COLUMN = _config['EXCEL']["estimate column"]
SPRINT_COLUMN = _config['EXCEL']["sprint column"]
START_ROW = int(_config['EXCEL']["start row"])
VALIDATION_COLOR = _config['EXCEL']["validation color"]
SPRINT_CUSTOM_FIELD = _config['SPRINT']["sprint custom field"]
SPRINT_ID = int(_config['SPRINT']["sprint id"])
OPEN_STATUS = "2"


# CORE
jira = JIRA(server=JIRA_SERVER, basic_auth=(EMAIL, TOKEN))

is_excel_recent()
wb_obj = openpyxl.load_workbook(EXCEL_LOCATION)
worksheet = wb_obj.get_sheet_by_name(SPRITESHEET_NAME)

last_story = None
last_epic = None
last_epic_assignee = None
last_task = None

"""issue = jira.issue("LC-1746")
for field_name in issue.raw['fields']:
    if field_name == "customfield_10020":
        print("Field:", field_name, "Value:", issue.raw['fields'][field_name])"""

for row in range(START_ROW, worksheet.max_row+1):#worksheet.max_row+1
    #for column in "ADEF":  #Here you can add or reduce the columns
    cell_name = "{}{}".format(SUMMARY_COLUMN, row)
    summary_cell = worksheet[cell_name] # the value of the specific cell
    cell_name = "{}{}".format(ASSIGNEE_COLUMN, row)
    assignee_cell = worksheet[cell_name]

    if summary_cell.value is not None:
        color = f"_{summary_cell.fill.start_color.index}_"[1:-1]
        if color in VALIDATION_COLOR:
            cell_name = "{}{}".format(ISSUETYPE_COLUMN, row)
            issuetype_cell = worksheet[cell_name]
            if issuetype_cell.value == "Epic":
                cell_name = "{}{}".format(KEY_COLUMN, row)
                key_cell = worksheet[cell_name]
                if key_cell.value is not None:
                    last_epic = key_cell.value
                    last_epic_assignee = assignee_cell.value
                    print(summary_cell.value, last_epic_assignee, last_epic)

                last_story = None
                last_task = None

            elif issuetype_cell.value == "Story":
                cell_name = "{}{}".format(KEY_COLUMN, row)
                key_cell = worksheet[cell_name]
                if key_cell.value is not None:
                    last_story = key_cell.value
            elif issuetype_cell.value == "Task":
                cell_name = "{}{}".format(KEY_COLUMN, row)
                key_cell = worksheet[cell_name]
                if key_cell.value is not None:
                    last_task = key_cell.value

        color = f"_{summary_cell.fill.start_color.index}_"[1:-1]
        if color in VALIDATION_COLOR:
            cell_name = "{}{}".format(KEY_COLUMN, row)
            key_cell = worksheet[cell_name]
            if key_cell.value is None:
                cell_name = "{}{}".format(DESCRIPTION_COLUMN, row)
                description_cell = worksheet[cell_name]
                cell_name = "{}{}".format(PRIORITY_COLUMN, row)
                priority_cell = worksheet[cell_name]
                cell_name = "{}{}".format(ESTIMATE_COLUMN, row)
                estimate_cell = worksheet[cell_name]
                cell_name = "{}{}".format(SPRINT_COLUMN, row)
                sprint_cell = worksheet[cell_name]

                issuetype = issuetype_cell.value
                summary = summary_cell.value
                description = description_cell.value
                assignee = assignee_cell.value
                priority = priority_cell.value
                estimate = estimate_cell.value
                sprint_customfield = {"number": SPRINT_CUSTOM_FIELD, "content": SPRINT_ID}
                customfield = sprint_customfield
                subtask_customfield = None
                sprint = sprint_cell.value

                if summary is not None:
                    summary = f"{PREFIX}{summary}"

                if sprint == "Backlog":
                    if assignee is None:
                        assignee = "Unassigned"
                    customfield = None

                if issuetype is not None and assignee is not None:
                    if issuetype == "Epic":
                        new_issue = create_jira(PROJECT, issuetype, summary,
                                                description=description, assignee=assignee,
                                                priority=priority, reporter=assignee, customfield=customfield)

                        last_epic = str(new_issue)
                        last_story = None
                        last_task = None

                    elif issuetype == "Story":
                        new_issue = create_jira(PROJECT, issuetype, summary,
                                                description=description, assignee=assignee,
                                                priority=priority, parent=last_epic, reporter=last_epic_assignee,
                                                customfield=customfield)

                        last_story = str(new_issue)

                    elif issuetype == "Task":
                        if priority is not None:

                            if last_epic is not None:
                                new_issue = create_jira(PROJECT, issuetype, summary,
                                                        description=description, assignee=assignee,
                                                        priority=priority, parent=last_epic, estimate=estimate,
                                                        reporter=last_epic_assignee, customfield=customfield)
                            else:
                                new_issue = create_jira(PROJECT, issuetype_cell.value, summary,
                                                        description=description, assignee=assignee,
                                                        priority=priority, estimate=estimate, customfield=customfield)
                            if last_story is not None:
                                jira.create_issue_link(
                                    type = "Relates",
                                    inwardIssue = last_story,
                                    outwardIssue = str(new_issue)
                                )
                            last_task = key_cell.value
                        else:
                            print(f"row {row} has incomplete definition")

                    elif issuetype == "Sub-task":

                        if priority is not None:
                            if last_task is not None:
                                new_issue = create_jira(PROJECT, issuetype, summary,
                                                        description=description, assignee=assignee,
                                                        priority=priority, parent=last_task, estimate=estimate,
                                                        reporter=last_epic_assignee, customfield=subtask_customfield)
                            else:
                                new_issue = create_jira(PROJECT, issuetype, summary,
                                                        description=description, assignee=assignee,
                                                        priority=priority, estimate=estimate, customfield=subtask_customfield)
                            if last_story is not None:
                                jira.create_issue_link(
                                    type = "Relates",
                                    inwardIssue = last_story,
                                    outwardIssue = str(new_issue)
                                )
                        else:
                            print(f"row {row} has incomplete definition")

                    hyper_link = get_issue_hyperlink(new_issue)
                    key_cell.value = str(new_issue)
                    key_cell.hyperlink = hyper_link
                    key_cell.style = "Hyperlink"
                    print("Created issue: ", new_issue)
                    print(f"https://sgbcn.atlassian.net/browse/{new_issue}")
                    written = False
                    while written is False:
                        try:
                            wb_obj.save(EXCEL_LOCATION)
                            written = True
                        except:
                            print("Xlsx file seems to be open")
                            print("Press a Enter to continue")
                            input()
                else:
                    print(f"row {row} has incomplete definition")
                print("___________________________")
